#!/usr/bin/env python3
import openpyxl

def inspect_workbook(filename):
    """
    Inspects an Excel workbook and prints its sheet names, named ranges, and tables.
    """
    try:
        wb = openpyxl.load_workbook(filename, read_only=False) # read_only=True doesn't load tables in some versions
        print("Sheet names:")
        for sheet_name in wb.sheetnames:
            print(f"- {sheet_name}")
        
        print("\nNamed ranges:")
        for name in wb.defined_names:
             if hasattr(wb.defined_names[name], 'attr_text'):
                print(f"- {name}: {wb.defined_names[name].attr_text}")
             else:
                print(f"- {name}")

        print("\nTables:")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, 'tables'):
                for table_name in ws.tables.keys():
                    table = ws.tables[table_name]
                    print(f"- Table: {table_name} (Sheet: {sheet_name})")
                    print(f"  Range: {table.ref}")
                    if hasattr(table, 'tableColumns'):
                        columns = [col.name for col in table.tableColumns]
                        print(f"  Columns: {', '.join(columns)}")

    except Exception as e:
        print(f"Error inspecting workbook: {e}")

if __name__ == '__main__':
    inspect_workbook('Budget.xlsx')
