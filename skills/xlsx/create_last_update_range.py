#!/usr/bin/env python3
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

def setup_last_update_range(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        
        # Check if named range exists
        if 'LastRolloverUpdate' in wb.defined_names:
            print("Named range 'LastRolloverUpdate' already exists.")
            return

        if 'Rollovers' not in wb.sheetnames:
            print("Error: 'Rollovers' sheet not found.")
            return

        # Create named range on Rollovers!H1
        # Using specific scope if needed, but global is fine
        new_range = DefinedName('LastRolloverUpdate', attr_text='Rollovers!$H$1')
        wb.defined_names.add(new_range)
        
        # Initialize value if empty
        ws = wb['Rollovers']
        if not ws['H1'].value:
            ws['H1'] = "1900-01-01" # Default old date
            
        wb.save(filename)
        print("Created named range 'LastRolloverUpdate' at Rollovers!H1.")

    except Exception as e:
        print(f"Error setting up range: {e}")

if __name__ == '__main__':
    setup_last_update_range('Budget.xlsx')
