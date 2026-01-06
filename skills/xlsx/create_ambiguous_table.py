#!/usr/bin/env python3
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_ambiguous_table(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        
        if 'AmbiguousItems' in wb.sheetnames:
            print("Sheet 'AmbiguousItems' already exists.")
            return

        ws = wb.create_sheet(title="AmbiguousItems")
        
        # Headers
        headers = ['Item', 'IsAmbiguous', 'OverrideCount', 'Confidence']
        ws.append(headers)
        
        # Default Data
        data = [
            ['Walmart', 'TRUE', 0, 0],
            ['Amazon', 'TRUE', 0, 0],
            ['Costco', 'TRUE', 0, 0]
        ]
        
        for row in data:
            ws.append(row)
            
        # Create Table
        tab = Table(displayName="AmbiguousItems", ref="A1:D4")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        wb.save(filename)
        print("Created 'AmbiguousItems' sheet and table.")

    except Exception as e:
        print(f"Error creating table: {e}")

if __name__ == '__main__':
    create_ambiguous_table('Budget.xlsx')
